import excel2img
import openpyxl

signal = '采购业务员'


def last_row(wb, vendor: str):
    ws = wb['订单 ' + vendor]
    for cell in ws['A']:
        if cell.value is not None and isinstance(cell.value, str) and signal in cell.value:
            return cell.row - 1


filename = "bay 2.xlsx"

wb = openpyxl.load_workbook(filename)
vendors = [name.split(' ')[1] for name in wb.sheetnames if name.split(' ')[0] == '订单']
for vendor in vendors:
    print(vendor)
    row = last_row(wb, vendor)
    sheet = "'订单 " + vendor + "'!A2:M" + str(row)
    excel2img.export_img(filename, vendor + ".png", "", sheet)
