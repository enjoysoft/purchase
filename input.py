import os
import pandas as pd
import numpy as np
import glob
from openpyxl import load_workbook

base: str = 'data'
sample: str = '1101-1108美标件'
no_quote: str = '无法报价'
index_label: str = '序号'

signal: str = '合计'


def is_dollar(vendor: str):
    return vendor.encode('UTF-8').isalnum()


def is_hk(vendor: str):
    return vendor == 'JOHNSON'


def dollar_to_rmb(dollar: float):
    return dollar * 7 * 1.13 * 1.1 + 200


def dollar_to_hk(dollar: float):
    return dollar * 7 * 1.13 * 1.1 + 100


def last_row(ws):
    for cell in ws['A']:
        if cell.value is not None and isinstance(cell.value, str) and signal in cell.value:
            return cell.row - 1


def get_sub(a_dir):
    return [name for name in os.listdir(a_dir)
            if os.path.isdir(os.path.join(a_dir, name))]


def get_xls_file(a_dir):
    return glob.glob(base + "\\" + sample + "\\" + a_dir + "\\*.xlsx")[0]


header = ['序号', '产品名称', '型号规格', '规范', '单位', '数量', '品牌', '单价', '总价', '交货周期', '采购依据', '申请部门', '项目', 'MOQ']
shared_header = ['产品名称', '型号规格', '规范', '单位', '数量', '采购依据', '申请部门', '项目']
individual_header = ['品牌', '单价', '总价', '交货周期', 'MOQ']
assert len(shared_header) + len(individual_header) == len(header) - 1


def get_price(file: str, vendor: str):
    wb = load_workbook(filename=file, data_only=True)
    ws = wb['询价单']
    row = last_row(ws)

    # Read the cell values into a list of lists
    data_rows = []

    for row in ws['A16':'N' + str(row)]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Transform into data frame
    df = pd.DataFrame(data_rows, columns=header)
    df.set_index(index_label, inplace=True)
    df.replace({no_quote: np.nan, 0: np.nan, "": np.nan, None: np.nan, " ": np.nan, "N/A": np.nan}, inplace=True)
    return df


dirs = get_sub(base + "\\" + sample)
vendors = [s.split()[0] for s in dirs]
files = [get_xls_file(a_dir) for a_dir in dirs]
dfs = [get_price(file, vendor) for (file, vendor) in zip(files, vendors)]
all = pd.concat(dfs, keys=vendors)
hehe = all.unstack(level=0)
# TODO validation
res = hehe.xs(vendors[0], level=1, axis=1)[shared_header]

res3 = hehe[individual_header]
res3.to_excel(base + "\\" + "individual.xlsx")

res1 = hehe['总价'].astype('float64')
dollar_vendors = list(filter(is_dollar, vendors))

for vendor in vendors:
    if is_hk(vendor):
        res[vendor + '美元'] = res1[vendor]
        res1[vendor] = res1[vendor].apply(dollar_to_hk)
    elif is_dollar(vendor):
        res[vendor + '美元'] = res1[vendor]
        res1[vendor] = res1[vendor].apply(dollar_to_rmb)

mean = res1.mean(1)
mean.name = '平均'
std = res1.std(1)
std.name = '标准差'
count = res1.count(1)
count.name = '报价数'

selection = res1.idxmin(1, skipna=True)
selection.name = '供应商'
selection_quotes = res1.min(1)
selection_quotes.name = '报价'
res1 = res1.join(selection)
res1 = res1.join(selection_quotes)
res1 = res1.join(mean)
res1 = res1.join(std)
res1 = res1.join(count)

res2 = pd.concat([res, res1], axis=1)
res2.to_excel(base + "\\" + "quotes.xlsx")
