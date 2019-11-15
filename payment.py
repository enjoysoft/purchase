import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

d = pd.read_excel("payment/月度资金预算编制.xlsx", sheet_name='月度编制上传模版')
d.columns = d.columns.str.strip()
foo = d.groupby('备注')

book = load_workbook('payment/付款与挂账.xlsx')
detail = load_workbook('payment/预算分摊表.xlsx')
excel_writer = pd.ExcelWriter('payment/付款与挂账.xlsx', engine='openpyxl')
excel_writer.book = book
excel_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

detail_writer = pd.ExcelWriter('payment/预算分摊表.xlsx', engine='openpyxl')
detail_writer.book = detail
detail_writer.sheets = dict((ws.title, ws) for ws in detail.worksheets)

for item_id, items in foo:
    print('=============================')
    print(item_id)
    print('=============================')

    sheet_name = item_id
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])
    ws = book.copy_worksheet(book['挂账付款申请单'])
    ws.title = sheet_name

    assert not items.empty
    main_item = items.iloc[0]
    if len(items) == 1:
        ws['H6'] = main_item.备注
        ws['C7'] = main_item.合同乙方
        ws['H10'] = main_item.loc['预算金额（人民币）']
        ws['H27'] = main_item.预算科目编号
        ws['H28'] = main_item.预算项目编号
        ws['H29'] = main_item.预算部门编号  # 199
        ws['H30'] = main_item.IPT团队编号  # 99
        ws['H31'] = main_item.主合同编号  # 99
        ws['H32'] = main_item.资金预算订单号
        ws['H33'] = main_item.资金预算WBS编号
    else:
        if sheet_name in detail.sheetnames:
            detail.remove(detail[sheet_name])

        detail_ws = detail.copy_worksheet(detail['预算分摊表'])
        detail_ws.title = sheet_name
        detail_ws.insert_rows(3, len(items))
        excel_vendor_detail: pd.DataFrame = items[
            ['预算科目编号', '预算项目编号', '预算部门编号', 'IPT团队编号', '主合同编号', '资金预算订单号', '资金预算WBS编号', '预算金额（人民币）']]
        for r in dataframe_to_rows(excel_vendor_detail, index=False, header=False):
            detail_ws.append(r)

        acc = '见附件'
        ws['H6'] = main_item.备注
        ws['C7'] = main_item.合同乙方
        ws['H10'] = acc
        ws['H27'] = acc
        ws['H28'] = acc
        ws['H29'] = acc
        ws['H30'] = acc
        ws['H31'] = acc
        ws['H32'] = acc
        ws['H33'] = acc

excel_writer.save()
detail_writer.save()

