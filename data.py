import io

import docxtpl
import pandas as pd
import numpy as np
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from num2words import num2words
from rmb_upper import num2chn

selection_label: str = '供应商'
quote_label: str = '总价'
index_label: str = '序号'
formatting: str = '%.2f'
unit_label: str = '单价'
no_quote: str = '无法报价'
excel_filename: str = 'bay.xlsx'
contract: str = '???'

seed_order: int = 120
seed_contract: int = 33
pattern = {'BN': 'BN-2019-', 'C919': 'C919B-BN19', 'ARJ': 'ARJ21B-BN19', 'PB': 'PB-BN19', 'QB': 'QB-BN19'}

today: str = '2019年6月17日'
order: str = '0603-0611美标件'
start_day: str = '2019年6月3日'
end_day: str = '2019年6月11日'


def get_contract(cat: str):
    global seed_order, seed_contract
    if cat == 'BN':
        seed_order += 1
        return pattern[cat] + '{0:03}'.format(seed_order)
    else:
        seed_contract += 1
        return pattern[cat] + '{0:03}'.format(seed_contract)


def is_dollar(vname: str):
    return vname.encode('UTF-8').isalnum() or vname == "玥涵"


df: pd.DataFrame = pd.read_excel(excel_filename, sheet_name='quote', na_values=no_quote, index_col=0)
individual: pd.DataFrame = pd.read_excel("individual.xlsx", index_col=0, header=[0, 1])
vendors_raw: pd.Series = pd.read_excel(excel_filename, sheet_name='vendor', header=None, index_col=0, squeeze=True)
vendor_item_count = len(df)

vendors_name = vendors_raw.index
vendors = list(filter(lambda v: v in df, vendors_raw.index))
vendors_timeout: list = list(filter(lambda v: v not in df.columns, vendors_raw.index))
vendor_invalid = list(filter(lambda v: v not in vendors_raw.index, df.columns[8:]))
if vendor_invalid:
    raise ValueError("Invalid vendor name found! " + str(vendor_invalid))

raw_quotes = individual[quote_label]

quotes = df.loc[:, vendors]
selection = quotes.idxmin(1)
selection.name = selection_label
selection_quotes = quotes.min(1)
selection_quotes.name = quote_label
df = df.join(selection)
df = df.join(selection_quotes)

groups = [group for group in df.groupby(selection_label)]

quote_summary = pd.concat([raw_quotes.sum(), raw_quotes.count()], axis=1).sort_values(by=[1, 0],
                                                                                      ascending=[False, True])


def foo(raw_row: pd.Series):
    selection_item = selection[raw_row.name]
    if pd.isna(selection[raw_row.name]):
        return np.nan
    else:
        return raw_row[selection_item]


raw_selection_quotes = raw_quotes.apply(foo, axis=1)

df_raw = pd.concat([raw_selection_quotes, selection], axis=1)
raw_group = df_raw.groupby(selection_label)
raw_group_summary = pd.concat([raw_group.count(), raw_group.sum()], axis=1)
raw_group_summary.columns = [2, 3]
quote_summary = quote_summary.join(raw_group_summary)
detail: io.StringIO = io.StringIO()
for vendor, row in quote_summary.iterrows():
    if is_dollar(vendor):
        print("%s报价%d项，金额总计美元%.2f元，中标%d项，中标总金额$%.2f" % (
            vendors_raw[vendor], row.loc[1], row.loc[0], row.loc[2], row.loc[3]), sep='', file=detail)
    else:
        print("%s报价%d项，金额总计人民币%.2f元，中标%d项，中标总金额￥%.2f" % (
            vendors_raw[vendor], row.loc[1], row.loc[0], row.loc[2], row.loc[3]), sep='', file=detail)
for v in vendors_timeout:
    print(vendors_raw[v], '逾期未回复', sep='', file=detail)
detail_str: str = detail.getvalue()
detail.close()

vendor_all = len(vendors_raw)
vendor_count = len(quotes.columns[~quotes.isna().all()])

book = load_workbook(excel_filename)
excel_writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
excel_writer.book = book
excel_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

no_quotes = df.loc[df[selection_label].isna()].drop(vendors, 1).drop(selection_label, 1)
print('无报价')
print(no_quotes.to_csv(sep='\t', float_format=formatting))

for vendor_name, vendor_df in groups:
    print('=============================')
    print(vendor_name)
    print('=============================')

    vendor_product = vendor_df.drop(vendors, 1).drop(selection_label, 1)
    total_num = vendor_df[quote_label].sum()
    if total_num < 2e4:
        category = 'BN'
    else:
        project: str = vendor_df.groupby('项目')['总价'].sum().idxmax()
        if project == 'C919':
            category = 'C919'
        elif project == 'ARJ21':
            category = 'ARJ'
        elif project == '培训':
            category = 'PB'
        else:
            category = 'QB'

    contract = get_contract(category)

    vendor_individual = individual.xs(vendor_name, level=1, axis=1)
    vendor_product.drop(quote_label, 1, inplace=True)
    vendor_product = vendor_product.join(vendor_individual)
    total_real = vendor_product[quote_label].sum()
    hit = len(vendor_df.index)

    total_short: str = ('$' if is_dollar(vendor_name) else '￥') + formatting % total_real
    currency: str = ('美元' if is_dollar(vendor_name) else '人民币')
    total: str = currency + formatting % total_real

    print('总价', total)
    print(vendor_product.to_csv(sep='\t', float_format=formatting))

    numbers = ','.join(sorted(map(str, vendor_product['采购依据'].unique())))
    items = ','.join(sorted(vendor_product['项目'].unique()))

    vendor_quotes = vendor_df.loc[:, vendors]

    quote_groups = vendor_quotes.notna().groupby(vendors).groups
    quote_detail = [(v.tolist(), vendor_quotes.loc[v].sum(skipna=False).sort_values()) for k, v in
                    quote_groups.items()]

    # Word publish

    vendor_full_name = vendors_raw[vendor_name]

    doc = DocxTemplate("template.docx")

    context = {'numbers': numbers,
               'items': items,
               'vendor': vendor_full_name,
               'detail': docxtpl.R(detail_str),
               'total': total,
               'total_short': total_short,
               'vendor_all': vendor_all,
               'vendor_count': vendor_count,
               'contract': contract,
               'currency': currency,
               'order': order,
               'start_day': start_day,
               'end_day': end_day,
               'today': today,
               'hit': hit,
               'vendor_item_count': vendor_item_count,
               }
    doc.render(context)
    doc.save(vendor_name + ".docx")

    # Excel publish
    sheet_name = '订单 ' + vendor_name
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

    if is_dollar(vendor_name):
        ws = book.copy_worksheet(book['template_e'])
        ws.title = sheet_name
        ws['J6'] = vendor_full_name
        ws['J7'] = vendor_full_name
        ws['J4'] = numbers
        ws['J3'] = contract
        ws['I12'] = total_real
        ws['D13'] = num2words(total_real).upper()
    else:
        ws = book.copy_worksheet(book['template'])
        ws.title = sheet_name
        ws['J4'] = vendor_full_name
        ws['K3'] = numbers
        ws['k2'] = contract
        ws['I7'] = total_real
        ws['D8'] = "%s（￥%s）（含16%%增值税）" % (num2chn(total_real), total)
    columns = ['产品名称', '型号规格', '规范', '数量', '单位', '品牌', '单价', '总价', '交货周期', '采购依据', '申请部门', '项目', 'MOQ']
    excel_vendor_detail = vendor_product.reindex(columns=columns, fill_value='?')
    for r in dataframe_to_rows(excel_vendor_detail, index=True, header=False):
        ws.append(r)

excel_writer.save()
